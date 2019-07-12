from selenium import webdriver
from selenium.webdriver.support.ui import Select
from time import sleep
import openpyxl

s = 'ko' #input('과목(국어: ko, 영어: eng, 수학: math, 과학: sci):')
ms = [3,4,6,7,9,10,11]

write_wb = openpyxl.load_workbook("ko.xlsx")

 #원 숫자 변환 딕셔너리
roundToN = {'①':1,'②':2,'③':3,'④':4,'⑤':5}

#크롬 실행
driver = webdriver.Chrome('chromedriver.exe')

#3초 대기
driver.implicitly_wait(3)

#EBSi 로그인 페이지 접속
driver.get('http://www.ebsi.co.kr/ebs/pot/potl/login.ebs?destination=/index.jsp&alertYn=N')

#아이디, 비밀번호 입력
driver.find_element_by_name('userid').send_keys('encoder0511')
driver.find_element_by_name('passwd').send_keys('paekun0511')

#로그인 버튼 클릭
driver.find_element_by_class_name('loginWrap').find_element_by_tag_name('button').click()

#기출문제 페이지 접속
driver.get('http://www.ebsi.co.kr/ebs/xip/xipc/previousPaperList.ebs')

for y in range(2016, 2017):
    write_ws = write_wb.create_sheet(str(y))
    #Select 박스에서 연도 선택
    select = Select(driver.find_element_by_id('yearNum'))
    select.select_by_value(str(y))   
    driver.find_element_by_name('mAllChk').click()
    for m in ms:
        if m != 6:
            #체크 박스에서 월 선택
            driver.find_element_by_id('m'+str(m)).click()
            sleep(2)
            #ko, math, eng, koh, sci
            driver.find_element_by_id('subTab_'+s).click()
            sub=1
            sub = int(sub)
            scr = driver.find_elements_by_tag_name('tr')[sub].find_element_by_class_name('btn_apply').get_attribute('href')
            driver.execute_script(scr)
            driver.find_element_by_id('m'+str(m)).click()
            sleep(3)
            driver.switch_to.window(driver.window_handles[-1]) #창 전환
            driver.find_element_by_id('mode2').click()
            study = driver.find_element_by_class_name('study-mode')
            for num in range(1, 46):
                #결과
                result = roundToN[study.find_element_by_id('boardcorrect_'+str(num)).find_element_by_class_name('red').get_attribute("textContent")[0]]

                #답지
                txt = []
                for text in study.find_element_by_id('Explanation_'+str(num)).find_elements_by_css_selector('span'):
                    txt.append(text.get_attribute("textContent"))
                write_ws.cell(num, (2*(ms.index(m)+1))-1, result)
                write_ws.cell(num, (2*(ms.index(m)+1)), "\n".join(txt))
            driver.close()
            driver.switch_to.window(driver.window_handles[-1]) #창 전환
            write_wb.save('ko.xlsx')










