import openpyxl

class FindResult:
    def __init__(self, year, month, num, result=False):
        this.year = year
        this.month = month
        this.num = num
        this.result = result


def main():
    months = {'3':1,'4':2,'6':3,'7':4,'9':5,'10':6,'수능':7, '11':8}

    wb = openpyxl.load_workbook("ko.xlsx")

    year = str(2019)

    m = '3'

    num = 45
    
    result = False
    
    if result:
        print(m+'월 모의고사 '+str(num)+'번 정답은'+ str(wb[year].cell(num,2*months[m]-1).value)+'입니다.')
        print('해설을 읽어드리겠습니다.')
        print(wb[year].cell(num,2*n).value)
    elif not result:
        print(m+'월 모의고사 '+str(num)+'번 정답은'+ str(wb[year].cell(num,2*months[m]-1).value)+'입니다.')
    

if __name__ == '__main__':
	main()